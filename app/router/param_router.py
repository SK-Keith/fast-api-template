#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
@Project ：fast-api-template
@File    ：cookie_router.py
@Author  ：Keith007
@Date    ：2024/3/27 22:29
"""
import os
from typing import List, Optional
from fastapi import APIRouter, Cookie, Request, Header, Form, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.responses import JSONResponse
from app.utils.log import logger
from app.types import response
import aiofiles
import openpyxl
import asyncio
import re
from bs4 import BeautifulSoup
import json
import subprocess
import time
import concurrent.futures


router = APIRouter(prefix="/param", tags=["更多参数接收示例"])

@router.get("/container/ip")
async def pathParamReceive2():
    return {
        # "ip": "47.107.103.126",
        "ip": "127.0.0.1"
    }

@router.get("/cookie/key", summary="接收cookie中指定的key")
async def cookieKey(user_name: Optional[str] = Cookie(None)):
    """接收cookie中指定的key"""
    return {"user_name": user_name}


@router.get("/cookie/all", summary="所有cookie值")
async def cookieParams(request: Request):
    """接收cookie值"""
    return {"cookies": request.cookies}


@router.get("/header/key")
async def headerKey(x_platform: Optional[str] = Header(None)):
    """ 从header中获取指定key"""
    return {"x_platform": x_platform}


@router.get("/header/keys")
async def headerKey(x_ip: Optional[List[str]] = Header(None)):
    """ 从header中获取重复key的值"""
    return {"x_ip": x_ip}


@router.post("/form/key")
async def formKey(username: str = Form(...), password: str = Form(...)) -> response.HttpResponse:
    """ 接收表单中的参数"""
    body = {
        "username": username,
        "password": password
    }
    return response.ResponseSuccess(body)


@router.post("/upload/file")
async def uploadFile(file: UploadFile = Form(...), fileType: str = Form(...)) -> response.HttpResponse:
    """ 文件上传"""
    try:
        # 构造保存目录
        save_path = os.path.join(os.getcwd(), "tmp", fileType)
        # 不存在则创建目录
        os.makedirs(save_path, exist_ok=True)
        # 拼接文件全路径
        file_path = os.path.join(save_path, file.filename)
        # 读取文件内容并写入目标文件
        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)
        body = {
            "fileName": file.filename,
            "fileType": fileType,
            "size": file.size,
        }
        return response.ResponseSuccess(body)
    except Exception as e:
        return response.ResponseFail("文件上传失败:" + str(e))


@router.post("/file/status")
async def get_file_status(data: dict) -> JSONResponse:
    logger.info("This is file status")
    try:
        # 从请求参数中获取文件列表
        keywords = data.get("keywords", [])

        # 遍历关键字列表，检查每个关键字对应的文件是否存在于根目录中
        status_list = []
        for keyword in keywords:
            file_path = os.path.join(os.getcwd(), f"{keyword}.xlsx")
            # file_path = f"/{keyword}.xlsx"  # 假设文件都是在根目录下
            print('file_path', file_path)
            if os.path.isfile(file_path):
                status_list.append({"keyword": keyword, "status": 1})
            else:
                status_list.append({"keyword": keyword, "status": 0})

        # 返回状态列表
        return response.ResponseSuccess(status_list)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/file/download")
async def downloadFile(keyword: str) -> FileResponse:
    """ 文件下载 """
    try:
        # 下载 Excel 文件
        print('keyword:', keyword)
        file_path = await download_excel(keyword)
        print('file_path:', file_path)
        if file_path:
            return FileResponse(file_path, filename=f"{keyword}.xlsx")
        else:
            return "数据还未爬取完成"
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件下载失败: {str(e)}")

@router.post("/upload/fileAndDownload")
async def uploadFile(file: UploadFile = File(...)) -> response.HttpResponse:
    """ 文件上传并解析 """
    print('开始上传')
    try:
        # 提取文件类型
        file_extension = file.filename.split(".")[-1]
        if file_extension != "xlsx":
            raise HTTPException(status_code=400, detail=f"只允许上传 xlsx 类型的文件")

        # 保存文件到本地
        file_path = await save_uploaded_file(file)

        # 解析 Excel 文件
        data = parse_excel(file_path)

        # 执行数据分析
        # await analyseData(data)
        # 执行数据分析，不等待其完成
        analyseDataTask = asyncio.create_task(analyseData(data))

        return response.ResponseSuccess(data)
    except Exception as e:
        print("上传文件时出现异常:", e)
        # raise HTTPException(status_code=500, detail=f"文件上传失败: {str(e)}")


# 保存上传文件到本地临时目录
async def save_uploaded_file(file: UploadFile) -> str:
    # 构造保存目录
    save_path = os.path.join(os.getcwd(), "tmp")
    # 不存在则创建目录
    os.makedirs(save_path, exist_ok=True)

    # 拼接文件全路径
    file_path = os.path.join(save_path, file.filename)

    # 保存文件到本地
    with open(file_path, "wb") as f:
        contents = await file.read()
        f.write(contents)

    return file_path


# 解析 Excel 文件并返回数据
# def parse_excel(file_path: str) -> list[list[str]]:
#     # 打开 Excel 文件并解析数据
#     workbook = openpyxl.load_workbook(file_path)
#     sheet = workbook.active
#     data = []
#     for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
#         page, keyword = row
#         data.append([page, keyword])
#
#     return data

def parse_excel(file_path: str) -> list[list[str]]:
    # 打开 Excel 文件并解析数据
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        page, keyword = row
        # 将关键词中的空格替换为 "+"
        keyword = keyword.replace(" ", "+")
        data.append([page, keyword, 0])

    return data


async def analyseData(data):
    print('analyseData:', data)
    start_time = time.time()  # 记录开始时间
    print("第一步：根据excel中的分页数和关键词爬取前台")
    parse_front(data)
    # 获取当前目录
    current_dir = os.getcwd()
    # 构建 htmlFolder 文件夹的完整路径
    directory = os.path.join(current_dir, 'tmp')
    loop = asyncio.get_event_loop()
    for pair in data:
        concurrency_limit, keyword, index = pair
        data = await loop.run_in_executor(None, lambda: extract_data_from_html(directory, keyword))
        print(f"analyseData concurrency_limit:：{concurrency_limit} keyword：{keyword} 秒")
        asins = set()
        for row in data[1:]:
            asin = row[0]
            if asin:
                files = os.listdir(directory)
                for file in files:
                    if asin in file:
                        break
                else:
                    asins.add(asin)

        asins_string = ','.join(asins)
        print("第二步：根据爬取出来的bought有数据的asin进行爬取二级页面")
        start_time_step2 = time.time()
        await loop.run_in_executor(None, lambda: parse_detail(asins_string))
        end_time_step2 = time.time()
        time_taken_step2 = end_time_step2 - start_time_step2

        print("第三步：解析并输出到excel")
        start_time_step3 = time.time()
        process_data(data, directory)
        # encoded_keyword = quote_plus(keyword)
        # output_file = os.path.join(directory, encoded_keyword + ".xlsx")
        # output_file = os.path.join(directory, keyword + ".xlsx")
        output_file = keyword + ".xlsx"
        print('output_file111:', output_file)
        output_file = keyword + ".xlsx"
        write_to_excel(data, output_file)
        end_time_step3 = time.time()
        time_taken_step3 = end_time_step3 - start_time_step3
        print(f"第二步耗时：{time_taken_step2:.2f} 秒，第三步耗时：{time_taken_step3:.2f} 秒")

    end_time = time.time()
    total_time = end_time - start_time
    print(f"总耗时：{total_time:.2f} 秒")

def extract_data_from_html(directory, keyword):
    data = [
        ["ASIN", "数量", "标题", "价格", "sponsored", "类目", "Material", "Brand", "Color", "Item Weight", "Product Dimensions",
         "Country of Origin", "Customer Reviews", "Date First Available"]]

    for filename in os.listdir(directory):
        # 检查文件名是否以 .html 结尾，并且包含关键字
        if filename.endswith('.html') and keyword in filename:
            filepath = os.path.join(directory, filename)
            print('文件名：' + filepath)
            with open(filepath, 'r', encoding='utf-8') as file:
                html_content = file.read()

            soup = BeautifulSoup(html_content, 'html.parser')

            items = soup.select('div[data-asin]')
            # 打印标题和对应的数量
            for item in items:
                asin = item.get('data-asin')
                title = item.find(class_='a-size-base-plus a-color-base a-text-normal')
                quantity = item.find(class_='a-size-base a-color-secondary')
                price = item.find(class_='a-size-base a-color-secondary')
                sponsored = item.find(class_='aok-inline-block puis-sponsored-label-info-icon')
                # 数量
                quantity_text = ''
                if quantity:
                    quantity_text = quantity.get_text()
                else:
                    if asin and not quantity_text:
                        item_text = item.get_text()
                        boughtKeyword = ' bought in past month'
                        keyword_index = item_text.find(boughtKeyword)
                        if keyword_index != -1:
                            quantity_text = item_text[:keyword_index].strip()
                            if quantity_text[0].isdigit():
                                print("数量1:", quantity_text)
                            else:
                                print("数量1: I don't know")

                if not quantity_text or not re.search(r'\d',
                                                      quantity_text) or not asin or "bought" not in quantity_text:
                    # 还要判断是不包含了sponsored
                    if not sponsored:
                        continue

                # 标题
                title_text = ''
                if title:
                    title_text = title.get_text()
                else:
                    # 如果title为空但是item包含span标签，则尝试从span标签中获取标题
                    span_title = item.find('span', class_='a-size-medium a-color-base a-text-normal')
                    if span_title:
                        title_text = span_title.get_text()
                # 价格 <span class="a-offscreen">$23.99</span>
                price = item.find('span', class_='a-offscreen')
                price_text = ''
                if price:
                    price_text = price.text.strip()
                # sponsored
                sponsored_text = ''
                if sponsored:
                    sponsored_text = 'sponsored'
                    if not asin:
                        link = item.find('a',
                                         class_='a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal')
                        href = link.get('href')
                        asin = extract_asin_from_link(href)

                row_data = [asin, quantity_text, title_text, price_text, sponsored_text]
                # 判断 data 中是否已经存在相同的 asin，如果不存在才添加新的数据行
                if asin not in [row[0] for row in data]:
                    data.append(row_data)

    return data


# 解析视频广告的ASIN
def extract_asin_from_link(link):
    match = re.search(r'/dp/(\w+)/', link)
    if match:
        return match.group(1)
    else:
        return None


def parse_detail(asins):
    # 如果asins为空，则不进行处理
    if not asins:
        print("ASIN列表为空，无需执行命令")
        return
    # 构建命令字符串
    script_path = os.path.join(os.getcwd(), "app", "newman", "run_newman_batch.js")
    print('script_path', script_path)
    # 构建命令字符串
    command = f'node "{script_path}" {asins}'
    # command = f'node run_newman_batch.js {asins}'
    print('command:', command)
    # 执行命令
    process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = process.communicate()

    # # 解析 stdout 中的响应数据
    # try:
    #     response_data = json.loads(stdout.decode('utf-8'))
    #     print('响应数据:', response_data)
    # except json.JSONDecodeError as e:
    #     print('解析响应数据时出错:', e)

    # 输出错误信息
    if stderr:
        print('错误信息:', stderr.decode('utf-8'))

    print('爬取二级完成！', asins)

def process_data(data, directory):
    # asins_to_process = set()  # 存储待处理的asin
    # for row in data[1:]:
    #     asin = row[0]
    #     if asin and asin not in asins_to_process:
    #         asins_to_process.add(asin)  # 将待处理的asin添加到集合中，以去重

    # 列出目录中的所有文件
    files = [filename for filename in os.listdir(directory) if filename.endswith('.html')]

    # 循环处理每个待处理的 ASIN
    parse_html_files_for_asins(directory, data, files)

def parse_html_files_for_asins(directory, data, files):
    # 定义数据表头顺序
    data_headers = ["ASIN", "数量", "标题", "价格", "sponsored", "类目", "Material", "Brand", "Color", "Item Weight",
                    "Product Dimensions",
                    "Country of Origin", "Customer Reviews", "Date First Available"]
    processed_files = set()  # 存储已处理过的文件名
    print('开始解析')
    start_time_step1 = time.time()
    for row in data[1:]:
        asin = row[0]
        while len(row) < 15:
            row.append('')
        file_found = False  # 标记是否找到符合条件的文件
        # 遍历文件，提取每个 ASIN 对应的 category 和 color
        try:
            for filename in files:
                if asin in filename and filename not in processed_files:
                    file_found = True  # 找到了符合条件的文件
                    filepath = os.path.join(directory, filename)
                    with open(filepath, 'r', encoding='utf-8') as file:
                        html_content = file.read()
                        soup = BeautifulSoup(html_content, 'html.parser')

                        # 解析html文件，获取category信息
                        div_element = soup.find('div', id='wayfinding-breadcrumbs_feature_div')
                        last_a_tag = div_element.find_all('a')[-1]
                        category = last_a_tag.get_text(strip=True)
                        # 将获取到的 category 存储起来
                        row[data_headers.index("类目")] = category


                        # 解析html文件，获取color信息
                        form_element = soup.find('form', id='twister')
                        if form_element:
                            label_element = form_element.find('label', class_='a-form-label')
                            if label_element:
                                label_text = label_element.get_text(strip=True)
                                if 'Color' in label_text:
                                    span_element = form_element.find('span', class_='selection')
                                    if span_element:
                                        color = span_element.get_text(strip=True)
                                        row[data_headers.index("Color")] = color

                        # 解析Brand、Color	Material
                        div_element = soup.find('div', class_='a-section a-spacing-small a-spacing-top-small')
                        if div_element:
                            # 找到包含产品细节的 table 元素
                            table_element = div_element.find('table', class_='a-normal a-spacing-micro')

                            if table_element:
                                # 遍历 table 元素中的每一行 tr
                                for tr_element in table_element.find_all('tr'):
                                    # 获取每一行中的 td 元素列表
                                    td_elements = tr_element.find_all('td')
                                    if len(td_elements) > 1:
                                        # 获取每个细节的标题和内容
                                        title_element = td_elements[0].find('span', class_='a-size-base a-text-bold')
                                        content_element = td_elements[1].find('span', class_='a-size-base po-break-word')
                                        if title_element and content_element:  # 检查元素是否存在
                                            detail_title = title_element.get_text(strip=True)
                                            detail_content = content_element.get_text(strip=True)
                                            if detail_title == 'Brand':
                                                row[data_headers.index("Brand")] = detail_content
                                            elif detail_title == 'Color':
                                                row[data_headers.index("Color")] = detail_content

                        # 解析Brand、Color
                        prod_details_element = soup.find('div', id='prodDetails')
                        if prod_details_element:
                            # 找到包含产品细节的表格元素
                            detail_table_element = prod_details_element.find('table', id='productDetails_detailBullets_sections1')
                            if detail_table_element:
                                # 遍历表格中的每个 tr 元素
                                for tr_element in detail_table_element.find_all('tr'):
                                    # 获取每个 tr 元素中的 th 和 td 元素
                                    th_element = tr_element.find('th',
                                                                 class_='a-color-secondary a-size-base prodDetSectionEntry')
                                    td_element = tr_element.find('td', class_='a-size-base prodDetAttrValue')
                                    if th_element and td_element:
                                        # 获取属性标题和内容
                                        detail_title = th_element.get_text(strip=True)
                                        detail_content = td_element.get_text(strip=True)
                                        if detail_title == 'Brand':
                                            row[data_headers.index("Brand")] = detail_content
                                        if detail_title == 'Material':
                                            row[data_headers.index("Material")] = detail_content
                                        elif detail_title == 'Color':
                                            row[data_headers.index("Color")] = detail_content
                                        elif detail_title == 'Item Weight':
                                            row[data_headers.index("Item Weight")] = detail_content
                                        elif detail_title == 'Color':
                                            row[data_headers.index("Color")] = detail_content
                                        elif detail_title == 'Product Dimensions':
                                            row[data_headers.index("Product Dimensions")] = detail_content
                                        elif detail_title == 'Country of Origin':
                                            row[data_headers.index("Country of Origin")] = detail_content
                                        elif detail_title == 'Date First Available':
                                            row[data_headers.index("Date First Available")] = detail_content
                                    # 解析 Customer Reviews
                                    average_reviews_element = tr_element.find('div', id='averageCustomerReviews')
                                    if average_reviews_element:
                                        # 找到包含平均评分的 span 元素
                                        average_rating_span = average_reviews_element.find('span',
                                                                                           class_='a-size-base a-color-base')
                                        if average_rating_span:
                                            average_rating = average_rating_span.get_text(strip=True)

                                        # 找到包含评分数量的 span 元素
                                        review_count_span = average_reviews_element.find('span', id='acrCustomerReviewText')
                                        if review_count_span:
                                            review_count = review_count_span.get_text(strip=True)

                                        # 找到包含总体评分的 span 元素
                                        overall_rating_span = average_reviews_element.find('span',
                                                                                           class_='reviewCountTextLinkedHistogram')
                                        if overall_rating_span:
                                            overall_rating_text = overall_rating_span['title']

                                        # 拼接字符串
                                        customer_reviews_str = f"{average_rating} ({review_count}, {overall_rating_text})"

                                        # 将结果添加到 row 中
                                        row[data_headers.index("Customer Reviews")] = customer_reviews_str
                    # 标记文件已处理
                    processed_files.add(filename)
                    break  # 找到了文件就退出内层循环
            if file_found:
                print(row)
                continue  # 如果找到了文件，则直接跳到下一个 row 的处理
        except Exception as e:
            print(f"An error occurred: {e}, asin:{asin}")
    end_time_step1 = time.time()
    time_taken_step1 = end_time_step1 - start_time_step1  # 计算第二步耗时
    print(f"解析完成耗时：{time_taken_step1}")

def write_to_excel(data, output_file):
    # 创建一个新的 Excel 文件
    wb = openpyxl.Workbook()
    # 选择默认的工作表
    ws = wb.active

    # 写入数据
    for row_data in data:
        ws.append(row_data)

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 20

    # 保存文件
    wb.save(output_file)

def parse_front(data):
    print('parse_front', data)
    # 打印数据
    for pair in data:
        try:
            # 构建命令字符串
            concurrency_limit, keyword, index = pair
            # 获取 run_newman_batch_pc.js 文件的绝对路径
            script_path = os.path.join(os.getcwd(), "app", "newman", "run_newman_batch_pc.js")
            print('script_path', script_path)
            # 构建命令字符串
            command = f'node "{script_path}" {concurrency_limit} {keyword}'
            # command = f'node run_newman_batch_pc.js {concurrency_limit} {keyword}'
            print('command:', command)
            # 执行命令
            process = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            stdout, stderr = process.communicate()

            # # 解析 stdout 中的响应数据
            # try:
            #     response_data = json.loads(stdout.decode('utf-8'))
            #     print('响应数据:', response_data)
            # except json.JSONDecodeError as e:
            #     print('解析响应数据时出错:', e)

            # 输出错误信息
            if stderr:
                print('错误信息:', stderr.decode('utf-8'))

            print('爬取前台完成！')
        except ValueError as e:
            # 长度不为 2 的情况
            print("parse_front:", e)
            continue

async def download_excel(keyword: str) -> str:
    """下载指定关键词的 Excel 文件"""
    # file_path = os.path.join(os.getcwd(), "tmp", f"{keyword}.xlsx")
    file_path = os.path.join(os.getcwd(), f"{keyword}.xlsx")
    print('file_path', file_path)
    print('yes:', os.path.isfile(file_path))
    if os.path.isfile(file_path):
        return file_path
    else:
        return None